import os
from openpyxl import load_workbook

nameData = '_Мониторинг поступления работ по актам приема передач.xlsx'  # Имя файла для сортировки
nameResult = 'nameResult.xlsx'
pathData = 'pathData'  # Путь к директории работы

rowData = 2  # Первая строка, с которой начинается сортировка в файле с данными
rowResult = 2  # Первая строка, с которой начинается ввод данных в конечном файле
sumObj = 0  # Счетчик одинковых строк

pagesResult = []  # Открытие буферного файла
wbResult = load_workbook('temp.xlsx')
pagesResult = wbResult.get_sheet_names()
sheetResult = wbResult[pagesResult[0]]

# os.chdir(folderPath)  # Переходим в директорию с файлом с данными

pagesData = []  # Открытие файла с данными
wbData = load_workbook(nameData)
pagesData = wbData.get_sheet_names()
sheetData = wbData[pagesData[0]]
# ===================================================================

counter = 0     # Счетчик строк при сравнении
exitCycle = 0   # Флаг на выход из цила сровнения

# Цикл по каждоый строчке
while sheetData['B' + str(rowData)].value is not None and sheetData['C' + str(rowData)].value is not None:
    if(sheetData['B' + str(rowData)].value is not '0000'):

        valueData_Name1 = sheetData['B' + str(rowData)].value
        valueData_Type1 = sheetData['C' + str(rowData)].value

        sumObj = 0  # Сброс счетчика количества совпадений
        counter = 0 # Сброс счетчика строк при сравнении
        exitCycle = 0  # Сброс флага на выход

        # Цикл сравнения каждой строчки с остальными
        while exitCycle is not 1:
            print('   По ', rowData, 'сравнение с ', rowData + counter)

            # Сохраняем значения следующей строки
            valueData_Name2 = sheetData['B' + str(rowData + counter)].value
            valueData_Type2 = sheetData['C' + str(rowData + counter)].value

            # Проверка на совпадение
            if valueData_Name1 == valueData_Name2 and valueData_Type1 == valueData_Type2 and valueData_Type2 is not '0000':
                print('Совпадение: ', valueData_Name1, '==', valueData_Name2, ' and \n', valueData_Type1, '==', valueData_Type2)

                # Замещаем совпадающую строку
                print('удаление строки: ', rowData + counter)
                sheetData['B' + str(rowData + counter)].value = '0000'
                sheetData['C' + str(rowData + counter)].value = '0000'

                sumObj += 1  # Увеличиваем счетчик совпадений

            # Проверка на окончание таблицы
            if sheetData['B' + str(rowData + counter)].value is None and sheetData['C' + str(rowData + counter)].value is None:
                print('===============================\n', 'Конец списка. Перемещение данных в строку ', rowResult)

                # Перемещение строки и количество объектов в итоговый файл
                sheetResult['A' + str(rowResult)].value = sheetData['A' + str(rowData)].value
                sheetResult['B' + str(rowResult)].value = sheetData['B' + str(rowData)].value
                sheetResult['C' + str(rowResult)].value = sheetData['C' + str(rowData)].value
                sheetResult['D' + str(rowResult)].value = sheetData['D' + str(rowData)].value
                sheetResult['G' + str(rowResult)].value = sumObj

                rowResult += 1  # Увеличение номера строки сравнения
                sumObj = 0  # Сброс количества совпадений
                exitCycle = 1   # Установить флаг на выход

            counter += 1  # увеличиваем номер строки примера

    rowData += 1
wbResult.save(nameResult)